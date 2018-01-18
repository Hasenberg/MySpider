#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import threading
import time
import re
from openpyxl import Workbook
from openpyxl import load_workbook

baseUrl = "https://book.douban.com"
headers = {"Host" : "book.douban.com",
"User-Agent" : "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0",
"Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
"Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
"Accept-Encoding": "gzip, deflate, br",
"Cookie": 'viewed="25955474_26923390_25782520_26319730_1418172_3988517_26919485_25663137_3137216_6109770"; _pk_ref.100001.3ac3=%5B%22%22%2C%22%22%2C1512526543%2C%22http%3A%2F%2Fwww.cnblogs.com%2Fvamei%2Farchive%2F2012%2F09%2F13%2F2682778.html%22%5D; _pk_id.100001.3ac3=3b92e6828d617456.1437973415.25.1512526543.1505279800.; gr_user_id=e6016067-2648-4c2a-bd3c-fa51ce9c2575; _vwo_uuid_v2=7D148D989CEE6417B89CE640EECE70C9|fcf27cd7f0217b9aff2ac7163a70529e; ll="108288"; bid=C3mHNREstWE; ap=1; dbcl2="144896259:wWPDQadbqss"; ps=y; push_noty_num=0; push_doumail_num=0; ck=dlhM',
"Connection": "keep-alive",
"Upgrade-Insecure-Requests": "1",
"Cache-Control": "max-age=0"
}
tagList = []
bookTypes = []
lock = threading.Lock()
xlsPath = "DoubanBook.xlsx"
proxy = {"https" : "https://120.86.89.131:9797"}
cookies = {'cookie': 'viewed="25955474_26923390_25782520_26319730_1418172_3988517_26919485_25663137_3137216_6109770"; _pk_ref.100001.3ac3=%5B%22%22%2C%22%22%2C1512526543%2C%22http%3A%2F%2Fwww.cnblogs.com%2Fvamei%2Farchive%2F2012%2F09%2F13%2F2682778.html%22%5D; _pk_id.100001.3ac3=3b92e6828d617456.1437973415.25.1512526543.1505279800.; gr_user_id=e6016067-2648-4c2a-bd3c-fa51ce9c2575; _vwo_uuid_v2=7D148D989CEE6417B89CE640EECE70C9|fcf27cd7f0217b9aff2ac7163a70529e; ll="108288"; bid=C3mHNREstWE; ap=1; dbcl2="144896259:wWPDQadbqss"; ps=y; push_noty_num=0; push_doumail_num=0; ck=dlhM'}

#获取标签种类
def getTags():
    print("读取标签种类。。。。。")
    curUlr = baseUrl + "/tag/?view=cloud"
    session = requests.session()
    response = session.get(curUlr, cookies=cookies, headers=headers, timeout=3)
    soup = BeautifulSoup(response.text, "lxml")
    table = soup.find("table", attrs={"class" : "tagCol"})
    for tr in table.findAll("tr"):
        for td in tr.findAll("td"):
            tag = td.find("a").text
            url = baseUrl + td.find("a")["href"]
            tagList.append((tag, url))

#每种标签下的书籍
def getTagBooks():
    while True:
        time.sleep(2)
        lock.acquire()
        if len(tagList) == 0:
            lock.release()
            break
        tag = tagList.pop()
        lock.release()
        url = tag[1]
        print(tag[0], "-------------------")
        response = requests.get(url, cookies=cookies, headers=headers)
        soup = BeautifulSoup(response.text, "lxml")

        #获取总页数
        pages = soup.find("div", attrs={"class" : "paginator"})
        pages = pages.findAll("a", text = re.compile("\d+"))
        pageNum = 0
        for page in pages:
            pageNum = max(pageNum, int(page.text))

        bookList = []
        #每页20项
        for index in range(0, pageNum):
            data = {"start" : index*20, "type" : "T"}
            response = requests.get(url, cookies=cookies, params=data, headers=headers)
            soup = BeautifulSoup(response.text, "lxml")

            ls = soup.find("ul", attrs={"class", "subject-list"})
            items = ls.findAll("li", attrs={"class" : "subject-item"})
            for item in items:
                info = item.find("div", attrs={"class" : "info"})
                try:
                    title = info.find("h2").find("a").text.strip()
                except:
                    title = "无"
                try:
                    line = info.find("div", attrs={"class" : "pub"}).text
                    line = line.split(" / ")
                except:
                    line = "无"
                try:
                    author = line[0].strip()
                except:
                    author = "无"
                try:
                    publish = line[1].strip()
                except:
                    publish = "无"
                try:
                    pubtime = line[2].strip()
                except:
                    pubtime = "无"
                try:
                    price = line[3].strip()
                except:
                    price = "无"
                try:
                    score = info.find("span", attrs={"class" : "rating_nums"}).text
                except:
                    score = "无"
                try:
                    commentNum = info.find("span", attrs={"class" : "pl"}).text.strip()
                    commentNum = re.sub("\D", "", commentNum)
                except:
                    commentNum = "无"
                bookList.append([author, title, publish, pubtime, price, score, commentNum])
                print(author, title, publish, pubtime, price, score, commentNum)
        saveBook(tag[0], bookList)
        bookTypes.append(bookList)

def saveBook(name, bookList):
    lock.acquire()
    try:
        wb = load_workbook(xlsPath)
    except:
        wb = Workbook()
        wb.save(xlsPath)
    ws = wb.create_sheet(name)
    ws.append(["作者", "书名", "出版社", "出版时间", "价格", "评分", "评论人数"])
    for book in bookList:
        ws.append(book)
    wb.save(xlsPath)
    lock.release()

getTags()
threadList = []
for i in range(15):
    t = threading.Thread(target=getTagBooks)
    t.start()
    t.join()
